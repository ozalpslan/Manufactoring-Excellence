from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.improvement import calculate_nvaa_savings


def create_powerbi_ready_dataset(oee_df: pd.DataFrame) -> pd.DataFrame:
    result = oee_df.copy()
    if result.empty:
        return result

    result["date"] = pd.to_datetime(result["date"], errors="coerce")
    result["date_key"] = result["date"].dt.strftime("%Y%m%d")
    result["year"] = result["date"].dt.year
    result["month"] = result["date"].dt.month
    result["month_name"] = result["date"].dt.strftime("%b")
    result["iso_week"] = result["date"].dt.isocalendar().week.astype("Int64")
    result["oee_percent"] = result["oee"] * 100
    result["availability_percent"] = result["availability"] * 100
    result["performance_percent"] = result["performance"] * 100
    result["quality_percent"] = result["quality"] * 100
    result["scrap_rate_percent"] = result["scrap_rate"] * 100
    result["downtime_rate_percent"] = result["downtime_rate"] * 100

    preferred_columns = [
        "date",
        "date_key",
        "year",
        "month",
        "month_name",
        "iso_week",
        "shift",
        "line",
        "machine",
        "product_family",
        "operator_team",
        "downtime_reason",
        "planned_shift_minutes",
        "planned_break_minutes",
        "planned_production_time_minutes",
        "unplanned_downtime_minutes",
        "operating_time_minutes",
        "ideal_cycle_time_seconds",
        "total_count",
        "defect_count",
        "good_count",
        "availability",
        "performance",
        "quality",
        "oee",
        "scrap_rate",
        "downtime_rate",
        "availability_percent",
        "performance_percent",
        "quality_percent",
        "oee_percent",
        "scrap_rate_percent",
        "downtime_rate_percent",
        "performance_over_100",
        "data_quality_status",
        "mes_correction_required",
        "mes_correction_reason",
        "source_row_id",
    ]
    columns = [column for column in preferred_columns if column in result.columns]
    remaining = [column for column in result.columns if column not in columns]
    return result[columns + remaining].sort_values(["date", "line", "shift"], ignore_index=True)


def _autosize_columns(writer: pd.ExcelWriter, sheet_name: str) -> None:
    worksheet = writer.sheets[sheet_name]
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


def _style_sheet(writer: pd.ExcelWriter, sheet_name: str) -> None:
    worksheet = writer.sheets[sheet_name]
    header_fill = PatternFill("solid", fgColor="1F6F78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    worksheet.freeze_panes = "A2"
    _autosize_columns(writer, sheet_name)


def create_excel_summary_report(
    oee_df: pd.DataFrame,
    manual_steps_df: pd.DataFrame,
    issues_df: pd.DataFrame,
    kaizen_df: pd.DataFrame,
) -> bytes:
    output = BytesIO()
    powerbi_df = create_powerbi_ready_dataset(oee_df)
    nvaa_df = calculate_nvaa_savings(manual_steps_df)

    summary = pd.DataFrame(
        [
            ["Overall OEE", oee_df["oee"].mean()],
            ["Availability", oee_df["availability"].mean()],
            ["Performance", oee_df["performance"].mean()],
            ["Quality", oee_df["quality"].mean()],
            ["Scrap Rate", oee_df["scrap_rate"].mean()],
            ["Total Downtime Minutes", oee_df["unplanned_downtime_minutes"].sum()],
            ["Monthly NVAA Hours Saved", nvaa_df["monthly_saved_hours"].sum()],
            ["Data Quality Issues", len(issues_df)],
        ],
        columns=["Metric", "Value"],
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary KPIs", index=False)
        powerbi_df.to_excel(writer, sheet_name="Power BI Dataset", index=False)
        issues_df.to_excel(writer, sheet_name="Data Quality Issues", index=False)
        kaizen_df.to_excel(writer, sheet_name="Kaizen Opportunities", index=False)
        nvaa_df.to_excel(writer, sheet_name="NVAA Savings", index=False)

        for sheet_name in writer.sheets:
            _style_sheet(writer, sheet_name)

    return output.getvalue()
